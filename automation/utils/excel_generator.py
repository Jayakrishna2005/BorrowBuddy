import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors for statuses
COLOR_PASS = "C6EFCE"  # Light Green
FONT_PASS = "006100"   # Dark Green
COLOR_FAIL = "FFC7CE"  # Light Red
FONT_FAIL = "9C0006"   # Dark Red
COLOR_SKIP = "FFEB9C"  # Light Yellow
FONT_SKIP = "9C6500"   # Dark Yellow
COLOR_HEADER = "4F81BD" # Slate Blue

def create_excel_report(report_path, test_cases, report_type="E2E"):
    """
    Generates a beautifully styled Excel report for test results.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Executed Test Cases
    ws_executed = wb.create_sheet(title="Executed Test Cases")
    # 2. Passed Tests
    ws_passed = wb.create_sheet(title="Passed Tests")
    # 3. Failed Tests
    ws_failed = wb.create_sheet(title="Failed Tests")
    # 4. Skipped Tests
    ws_skipped = wb.create_sheet(title="Skipped Tests")
    # 5. Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    # 6. Defect Summary
    ws_defects = wb.create_sheet(title="Defect Summary")

    headers = [
        "Test Case ID", "Module", "Priority", "Preconditions",
        "Test Steps", "Expected Result", "Actual Result", "Status", "Execution Time (ms)"
    ]

    # Helper function to style headers
    def apply_header_styles(ws):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

    # Style all raw result sheets
    apply_header_styles(ws_executed)
    apply_header_styles(ws_passed)
    apply_header_styles(ws_failed)
    apply_header_styles(ws_skipped)

    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    passed_count = 0
    failed_count = 0
    skipped_count = 0
    total_time = 0

    row_exec = 2
    row_pass = 2
    row_fail = 2
    row_skip = 2

    for tc in test_cases:
        status = tc.get("Status", "Skipped").upper()
        time_taken = tc.get("Execution Time (ms)", 0)
        total_time += time_taken

        # Populate rows
        row_data = [
            tc.get("Test Case ID", ""),
            tc.get("Module", ""),
            tc.get("Priority", "Medium"),
            tc.get("Preconditions", ""),
            tc.get("Test Steps", ""),
            tc.get("Expected Result", ""),
            tc.get("Actual Result", ""),
            status,
            time_taken
        ]

        # Add to Executed sheet
        for col_num, val in enumerate(row_data, 1):
            cell = ws_executed.cell(row=row_exec, column=col_num, value=val)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if col_num == 8: # Status column
                if status == "PASSED" or status == "PASS":
                    cell.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_PASS)
                elif status == "FAILED" or status == "FAIL":
                    cell.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_FAIL)
                else:
                    cell.fill = PatternFill(start_color=COLOR_SKIP, end_color=COLOR_SKIP, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_SKIP)
        row_exec += 1

        # Add to category specific sheet
        if status == "PASSED" or status == "PASS":
            passed_count += 1
            for col_num, val in enumerate(row_data, 1):
                cell = ws_passed.cell(row=row_pass, column=col_num, value=val)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                if col_num == 8:
                    cell.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_PASS)
            row_pass += 1
        elif status == "FAILED" or status == "FAIL":
            failed_count += 1
            for col_num, val in enumerate(row_data, 1):
                cell = ws_failed.cell(row=row_fail, column=col_num, value=val)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                if col_num == 8:
                    cell.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_FAIL)
            row_fail += 1
        else:
            skipped_count += 1
            for col_num, val in enumerate(row_data, 1):
                cell = ws_skipped.cell(row=row_skip, column=col_num, value=val)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                if col_num == 8:
                    cell.fill = PatternFill(start_color=COLOR_SKIP, end_color=COLOR_SKIP, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_SKIP)
            row_skip += 1

    # Auto-adjust column widths
    for ws in [ws_executed, ws_passed, ws_failed, ws_skipped]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # 5. Build Metrics Sheet
    ws_metrics.cell(row=1, column=1, value="Metric Name").font = Font(bold=True)
    ws_metrics.cell(row=1, column=2, value="Value").font = Font(bold=True)
    metrics = [
        ("Report Category", report_type),
        ("Total Executed Tests", passed_count + failed_count + skipped_count),
        ("Passed Tests", passed_count),
        ("Failed Tests", failed_count),
        ("Skipped Tests", skipped_count),
        ("Pass Percentage (%)", round((passed_count / (passed_count + failed_count + skipped_count) * 100), 2) if (passed_count + failed_count + skipped_count) > 0 else 0),
        ("Total Execution Time (s)", round(total_time / 1000, 2))
    ]
    for row_idx, (m_name, m_val) in enumerate(metrics, 2):
        ws_metrics.cell(row=row_idx, column=1, value=m_name).border = thin_border
        ws_metrics.cell(row=row_idx, column=2, value=m_val).border = thin_border
    ws_metrics.column_dimensions['A'].width = 30
    ws_metrics.column_dimensions['B'].width = 20

    # 6. Build Defects Summary Sheet
    ws_defects.cell(row=1, column=1, value="Defect ID").font = Font(bold=True)
    ws_defects.cell(row=1, column=2, value="Associated Test ID").font = Font(bold=True)
    ws_defects.cell(row=1, column=3, value="Module").font = Font(bold=True)
    ws_defects.cell(row=1, column=4, value="Error Message").font = Font(bold=True)
    ws_defects.cell(row=1, column=5, value="Severity").font = Font(bold=True)

    defect_row = 2
    for tc in test_cases:
        if tc.get("Status", "").upper() in ["FAILED", "FAIL"]:
            ws_defects.cell(row=defect_row, column=1, value=f"DEFECT-{tc.get('Test Case ID', '')}").border = thin_border
            ws_defects.cell(row=defect_row, column=2, value=tc.get("Test Case ID", "")).border = thin_border
            ws_defects.cell(row=defect_row, column=3, value=tc.get("Module", "")).border = thin_border
            ws_defects.cell(row=defect_row, column=4, value=tc.get("Actual Result", "Unknown Failure")).border = thin_border
            ws_defects.cell(row=defect_row, column=5, value=tc.get("Priority", "Medium")).border = thin_border
            defect_row += 1

    ws_defects.column_dimensions['A'].width = 15
    ws_defects.column_dimensions['B'].width = 25
    ws_defects.column_dimensions['C'].width = 20
    ws_defects.column_dimensions['D'].width = 50
    ws_defects.column_dimensions['E'].width = 15

    # Create directories if they don't exist
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    wb.save(report_path)
    print(f"Excel report saved to: {report_path}")
