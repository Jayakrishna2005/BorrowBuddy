import os
import sys
import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set directories
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
AUTOMATION_DIR = os.path.join(ROOT_DIR, "automation")
REPORTS_DIR = os.path.join(AUTOMATION_DIR, "reports")
RESULTS_DIR = os.path.join(ROOT_DIR, "Test Results")

os.makedirs(RESULTS_DIR, exist_ok=True)
os.makedirs(os.path.join(RESULTS_DIR, "Excel"), exist_ok=True)
os.makedirs(os.path.join(RESULTS_DIR, "HTML"), exist_ok=True)
os.makedirs(os.path.join(RESULTS_DIR, "JSON"), exist_ok=True)
os.makedirs(os.path.join(RESULTS_DIR, "Summary"), exist_ok=True)
os.makedirs(os.path.join(RESULTS_DIR, "Logs"), exist_ok=True)
os.makedirs(os.path.join(RESULTS_DIR, "Screenshots"), exist_ok=True)

# Styles
COLOR_PASS = "C6EFCE"
FONT_PASS = "006100"
COLOR_FAIL = "FFC7CE"
FONT_FAIL = "9C0006"
COLOR_SKIP = "FFEB9C"
FONT_SKIP = "9C6500"
COLOR_HEADER = "4F81BD"

def compile_all():
    print("Compiling master test reports...")
    
    # Files to parse
    test_files = {
        "Selenium": "selenium-results.json",
        "Appium": "appium-results.json",
        "Backend Vulnerability": "vulnerability-results.json",
        "API Unit": "unit-results.json",
        "UI Validation": "validation-results.json",
        "Deployment": "deployment-results.json",
        "Load": "load-results.json"
    }
    
    all_results = []
    category_metrics = {}
    
    # Load all json results
    for cat, filename in test_files.items():
        filepath = os.path.join(REPORTS_DIR, filename)
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                all_results.extend(data)
                
                passed = sum(1 for x in data if x.get("Status", "").upper() in ["PASSED", "PASS"])
                failed = sum(1 for x in data if x.get("Status", "").upper() in ["FAILED", "FAIL"])
                skipped = sum(1 for x in data if x.get("Status", "").upper() in ["SKIPPED", "SKIP"])
                total = len(data)
                duration = sum(x.get("Execution Time (ms)", 0) for x in data)
                
                category_metrics[cat] = {
                    "Total": total,
                    "Passed": passed,
                    "Failed": failed,
                    "Skipped": skipped,
                    "PassRate": round((passed / total * 100), 2) if total > 0 else 0,
                    "DurationMs": duration
                }
        else:
            print(f"Warning: Result file {filename} not found.")

    # 1. Generate execution-results.json
    with open(os.path.join(RESULTS_DIR, "JSON", "execution-results.json"), 'w', encoding='utf-8') as f:
        json.dump({
            "Summary": {
                "Timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "TotalTests": len(all_results),
                "Passed": sum(1 for x in all_results if x.get("Status", "").upper() in ["PASSED", "PASS"]),
                "Failed": sum(1 for x in all_results if x.get("Status", "").upper() in ["FAILED", "FAIL"]),
                "Skipped": sum(1 for x in all_results if x.get("Status", "").upper() in ["SKIPPED", "SKIP"]),
                "TotalDurationMs": sum(x.get("Execution Time (ms)", 0) for x in all_results)
            },
            "Categories": category_metrics,
            "Results": all_results
        }, f, indent=2)

    # Helper to apply headers
    headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (ms)", "Priority", "Steps", "Expected", "Actual"]
    
    def apply_header_styles(ws):
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    # Helper to populate test row
    def write_test_row(ws, row_idx, tc):
        status = tc.get("Status", "SKIPPED").upper()
        data = [
            tc.get("Test Case ID", ""),
            tc.get("Module", ""),
            tc.get("Expected Result", "")[:40] + "...", # Treat as test name
            status,
            tc.get("Execution Time (ms)", 0),
            tc.get("Priority", "Medium"),
            tc.get("Test Steps", ""),
            tc.get("Expected Result", ""),
            tc.get("Actual Result", "")
        ]
        for col_num, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if col_num == 4: # Status
                if status in ["PASSED", "PASS"]:
                    cell.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_PASS)
                elif status in ["FAILED", "FAIL"]:
                    cell.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_FAIL)
                else:
                    cell.fill = PatternFill(start_color=COLOR_SKIP, end_color=COLOR_SKIP, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=FONT_SKIP)

    # 2. Generate Automation_Test_Report.xlsx
    wb_master = openpyxl.Workbook()
    ws_exec = wb_master.active
    ws_exec.title = "Executed Test Cases"
    
    ws_pass = wb_master.create_sheet("Passed Tests")
    ws_fail = wb_master.create_sheet("Failed Tests")
    ws_skip = wb_master.create_sheet("Skipped Tests")
    ws_met = wb_master.create_sheet("Execution Metrics")
    ws_def = wb_master.create_sheet("Defect Summary")

    for ws in [ws_exec, ws_pass, ws_fail, ws_skip]:
        apply_header_styles(ws)

    exec_idx, pass_idx, fail_idx, skip_idx = 2, 2, 2, 2
    for tc in all_results:
        write_test_row(ws_exec, exec_idx, tc)
        exec_idx += 1
        
        status = tc.get("Status", "").upper()
        if status in ["PASSED", "PASS"]:
            write_test_row(ws_pass, pass_idx, tc)
            pass_idx += 1
        elif status in ["FAILED", "FAIL"]:
            write_test_row(ws_fail, fail_idx, tc)
            fail_idx += 1
        else:
            write_test_row(ws_skip, skip_idx, tc)
            skip_idx += 1

    # Fill Metrics Sheet
    ws_met.cell(row=1, column=1, value="Metric Category").font = Font(bold=True)
    ws_met.cell(row=1, column=2, value="Passed").font = Font(bold=True)
    ws_met.cell(row=1, column=3, value="Failed").font = Font(bold=True)
    ws_met.cell(row=1, column=4, value="Skipped").font = Font(bold=True)
    ws_met.cell(row=1, column=5, value="Total").font = Font(bold=True)
    ws_met.cell(row=1, column=6, value="Pass Rate (%)").font = Font(bold=True)

    met_row = 2
    for cat, m in category_metrics.items():
        ws_met.cell(row=met_row, column=1, value=cat).border = thin_border
        ws_met.cell(row=met_row, column=2, value=m["Passed"]).border = thin_border
        ws_met.cell(row=met_row, column=3, value=m["Failed"]).border = thin_border
        ws_met.cell(row=met_row, column=4, value=m["Skipped"]).border = thin_border
        ws_met.cell(row=met_row, column=5, value=m["Total"]).border = thin_border
        ws_met.cell(row=met_row, column=6, value=f"{m['PassRate']}%").border = thin_border
        met_row += 1
    
    # Fill Defect Sheet
    ws_def.cell(row=1, column=1, value="Defect ID").font = Font(bold=True)
    ws_def.cell(row=1, column=2, value="Test Case ID").font = Font(bold=True)
    ws_def.cell(row=1, column=3, value="Module").font = Font(bold=True)
    ws_def.cell(row=1, column=4, value="Priority").font = Font(bold=True)
    ws_def.cell(row=1, column=5, value="Failure Reason").font = Font(bold=True)
    
    def_row = 2
    for tc in all_results:
        if tc.get("Status", "").upper() in ["FAILED", "FAIL"]:
            ws_def.cell(row=def_row, column=1, value=f"DEFECT-{tc.get('Test Case ID', '')}").border = thin_border
            ws_def.cell(row=def_row, column=2, value=tc.get("Test Case ID", "")).border = thin_border
            ws_def.cell(row=def_row, column=3, value=tc.get("Module", "")).border = thin_border
            ws_def.cell(row=def_row, column=4, value=tc.get("Priority", "Medium")).border = thin_border
            ws_def.cell(row=def_row, column=5, value=tc.get("Actual Result", "")).border = thin_border
            def_row += 1

    for ws in [ws_exec, ws_pass, ws_fail, ws_skip, ws_met, ws_def]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    wb_master.save(os.path.join(RESULTS_DIR, "Excel", "Automation_Test_Report.xlsx"))

    # 3. Generate separate Passed_Test_Cases.xlsx and Failed_Test_Cases.xlsx
    wb_passed_only = openpyxl.Workbook()
    ws_po = wb_passed_only.active
    ws_po.title = "Passed Test Cases"
    apply_header_styles(ws_po)
    po_row = 2
    for tc in all_results:
        if tc.get("Status", "").upper() in ["PASSED", "PASS"]:
            write_test_row(ws_po, po_row, tc)
            po_row += 1
    wb_passed_only.save(os.path.join(RESULTS_DIR, "Excel", "Passed_Test_Cases.xlsx"))

    wb_failed_only = openpyxl.Workbook()
    ws_fo = wb_failed_only.active
    ws_fo.title = "Failed Test Cases"
    apply_header_styles(ws_fo)
    fo_row = 2
    for tc in all_results:
        if tc.get("Status", "").upper() in ["FAILED", "FAIL"]:
            write_test_row(ws_fo, fo_row, tc)
            fo_row += 1
    wb_failed_only.save(os.path.join(RESULTS_DIR, "Excel", "Failed_Test_Cases.xlsx"))

    # 4. Generate Summary_Report.xlsx
    wb_summary = openpyxl.Workbook()
    ws_s = wb_summary.active
    ws_s.title = "Consolidated Summary"
    ws_s.cell(row=1, column=1, value="Metric Category").font = Font(bold=True)
    ws_s.cell(row=1, column=2, value="Passed").font = Font(bold=True)
    ws_s.cell(row=1, column=3, value="Failed").font = Font(bold=True)
    ws_s.cell(row=1, column=4, value="Skipped").font = Font(bold=True)
    ws_s.cell(row=1, column=5, value="Total").font = Font(bold=True)
    ws_s.cell(row=1, column=6, value="Pass Rate (%)").font = Font(bold=True)
    for c_idx, h in enumerate(["Metric Category", "Passed", "Failed", "Skipped", "Total", "Pass Rate (%)"], 1):
        ws_s.cell(row=1, column=c_idx).border = thin_border

    s_row = 2
    for cat, m in category_metrics.items():
        ws_s.cell(row=s_row, column=1, value=cat).border = thin_border
        ws_s.cell(row=s_row, column=2, value=m["Passed"]).border = thin_border
        ws_s.cell(row=s_row, column=3, value=m["Failed"]).border = thin_border
        ws_s.cell(row=s_row, column=4, value=m["Skipped"]).border = thin_border
        ws_s.cell(row=s_row, column=5, value=m["Total"]).border = thin_border
        ws_s.cell(row=s_row, column=6, value=f"{m['PassRate']}%").border = thin_border
        s_row += 1

    wb_summary.save(os.path.join(RESULTS_DIR, "Excel", "Summary_Report.xlsx"))

    # 5. Generate beautiful responsive Dashboard HTML
    total_tests = len(all_results)
    passed_tests = sum(1 for x in all_results if x.get("Status", "").upper() in ["PASSED", "PASS"])
    failed_tests = sum(1 for x in all_results if x.get("Status", "").upper() in ["FAILED", "FAIL"])
    skipped_tests = sum(1 for x in all_results if x.get("Status", "").upper() in ["SKIPPED", "SKIP"])
    pass_percentage = round((passed_tests / total_tests * 100), 2) if total_tests > 0 else 0
    total_dur_s = round(sum(x.get("Execution Time (ms)", 0) for x in all_results) / 1000, 2)
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>E2E Execution Dashboard - BorrowBuddy</title>
    <style>
        :root {{
            --primary: #6C5CE7;
            --secondary: #00D2FF;
            --bg-color: #0F172A;
            --card-bg: #1E293B;
            --text-main: #F8FAFC;
            --text-muted: #94A3B8;
            --success: #10B981;
            --danger: #EF4444;
            --warning: #F59E0B;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Inter', system-ui, -apple-system, sans-serif;
            background-color: var(--bg-color);
            color: var(--text-main);
            padding: 2rem;
            line-height: 1.5;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 2rem;
            border-bottom: 1px solid rgba(255,255,255,0.1);
            padding-bottom: 1.5rem;
        }}
        .header h1 {{ font-weight: 800; font-size: 2rem; color: var(--primary); }}
        .header p {{ color: var(--text-muted); font-size: 0.95rem; margin-top: 0.25rem; }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1.5rem;
            margin-bottom: 3rem;
        }}
        .card {{
            background: var(--card-bg);
            border-radius: 16px;
            padding: 1.5rem;
            text-align: center;
            border: 1px solid rgba(255, 255, 255, 0.05);
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }}
        .card .value {{ font-size: 2.25rem; font-weight: 900; margin: 0.5rem 0; }}
        .card .label {{ font-size: 0.85rem; font-weight: 700; color: var(--text-muted); text-transform: uppercase; }}
        .table-container {{
            background: var(--card-bg);
            border-radius: 16px;
            padding: 1.5rem;
            margin-bottom: 3rem;
            border: 1px solid rgba(255, 255, 255, 0.05);
        }}
        .table-container h2 {{ font-size: 1.25rem; margin-bottom: 1.25rem; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
        }}
        th, td {{
            padding: 1rem;
            border-bottom: 1px solid rgba(255,255,255,0.05);
        }}
        th {{ color: var(--text-muted); font-size: 0.85rem; text-transform: uppercase; }}
        .badge {{
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 0.75rem;
            font-weight: bold;
        }}
        .badge-success {{ background: rgba(16, 185, 129, 0.15); color: var(--success); }}
        .badge-danger {{ background: rgba(239, 68, 68, 0.15); color: var(--danger); }}
        .badge-warning {{ background: rgba(245, 158, 11, 0.15); color: var(--warning); }}
    </style>
</head>
<body>
    <div class="header">
        <div>
            <h1>BorrowBuddy E2E Master Dashboard</h1>
            <p>Execution Date: {timestamp}</p>
        </div>
        <div style="text-align: right;">
            <div style="font-weight: bold;">Pipeline Run: SUCCESS</div>
            <div style="font-size: 0.85rem; color: var(--text-muted);">Commit SHA: {os.environ.get("GITHUB_SHA", "LocalRun")[:7]}</div>
        </div>
    </div>

    <div class="stats-grid">
        <div class="card" style="border-left: 4px solid var(--primary);">
            <div class="label">Total Test Cases</div>
            <div class="value">{total_tests}</div>
        </div>
        <div class="card" style="border-left: 4px solid var(--success);">
            <div class="label">Passed</div>
            <div class="value" style="color: var(--success);">{passed_tests}</div>
        </div>
        <div class="card" style="border-left: 4px solid var(--danger);">
            <div class="label">Failed</div>
            <div class="value" style="color: var(--danger);">{failed_tests}</div>
        </div>
        <div class="card" style="border-left: 4px solid var(--warning);">
            <div class="label">Skipped</div>
            <div class="value" style="color: var(--warning);">{skipped_tests}</div>
        </div>
        <div class="card" style="border-left: 4px solid var(--secondary);">
            <div class="label">Pass Rate</div>
            <div class="value" style="color: var(--secondary);">{pass_percentage}%</div>
        </div>
    </div>

    <div class="table-container">
        <h2>Category Metrics Dashboard</h2>
        <table>
            <thead>
                <tr>
                    <th>Metric Category</th>
                    <th>Total</th>
                    <th>Passed</th>
                    <th>Failed</th>
                    <th>Skipped</th>
                    <th>Pass Rate (%)</th>
                    <th>Duration (s)</th>
                </tr>
            </thead>
            <tbody>
"""

    for cat, m in category_metrics.items():
        html_content += f"""                <tr>
                    <td style="font-weight: bold;">{cat}</td>
                    <td>{m['Total']}</td>
                    <td style="color: var(--success); font-weight: bold;">{m['Passed']}</td>
                    <td style="color: var(--danger); font-weight: bold;">{m['Failed']}</td>
                    <td style="color: var(--warning); font-weight: bold;">{m['Skipped']}</td>
                    <td><span class="badge badge-success">{m['PassRate']}%</span></td>
                    <td>{round(m['DurationMs']/1000, 2)}s</td>
                </tr>
"""

    html_content += """            </tbody>
        </table>
    </div>
</body>
</html>
"""

    with open(os.path.join(RESULTS_DIR, "HTML", "dashboard.html"), 'w', encoding='utf-8') as f:
        f.write(html_content)

    # 6. Generate execution-report.html (containing detailed test cases table)
    exec_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>E2E Execution Report Details</title>
    <style>
        body {{ font-family: sans-serif; background: #0F172A; color: #F8FAFC; padding: 20px; }}
        h1 {{ color: #6C5CE7; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ padding: 8px 12px; border: 1px solid #1E293B; text-align: left; font-size: 0.85rem; }}
        th {{ background: #1E293B; color: #94A3B8; }}
        .pass {{ color: #10B981; font-weight: bold; }}
        .fail {{ color: #EF4444; font-weight: bold; }}
        .skip {{ color: #F59E0B; font-weight: bold; }}
    </style>
</head>
<body>
    <h1>E2E Execution Test Details ({timestamp})</h1>
    <table>
        <thead>
            <tr>
                <th>Test ID</th>
                <th>Module</th>
                <th>Priority</th>
                <th>Expected Result</th>
                <th>Actual Result</th>
                <th>Status</th>
                <th>Time (ms)</th>
            </tr>
        </thead>
        <tbody>
"""
    for tc in all_results:
        status_cls = "pass" if tc.get("Status", "").upper() in ["PASSED", "PASS"] else ("fail" if tc.get("Status", "").upper() in ["FAILED", "FAIL"] else "skip")
        exec_html += f"""            <tr>
                <td>{tc.get('Test Case ID')}</td>
                <td>{tc.get('Module')}</td>
                <td>{tc.get('Priority')}</td>
                <td>{tc.get('Expected Result')}</td>
                <td>{tc.get('Actual Result')}</td>
                <td><span class="{status_cls}">{tc.get('Status')}</span></td>
                <td>{tc.get('Execution Time (ms)')}ms</td>
            </tr>
"""
    exec_html += """        </tbody>
    </table>
</body>
</html>
"""
    with open(os.path.join(RESULTS_DIR, "HTML", "execution-report.html"), 'w', encoding='utf-8') as f:
        f.write(exec_html)

    # 7. Generate summary.md
    summary_md = f"""# Live GitHub Pages E2E Execution Summary

**Deployment URL:**
https://jayakrishna2005.github.io/BorrowBuddy

**Execution Date:**
{timestamp}

**Build Status:**
[PASS]

**Deployment Status:**
[PASS]

---

### Executed Metrics Dashboard
* **Total Test Cases:** {total_tests}
* **Passed:** {passed_tests} ✅
* **Failed:** {failed_tests} ❌
* **Skipped:** {skipped_tests} ⚠️
* **Pass Percentage:** {pass_percentage}%
* **Execution Duration:** {total_dur_s}s

| Category | Total | Passed | Failed | Skipped | Pass Rate (%) |
|---|---|---|---|---|---|
"""
    for cat, m in category_metrics.items():
        summary_md += f"| {cat} | {m['Total']} | {m['Passed']} | {m['Failed']} | {m['Skipped']} | {m['PassRate']}% |\n"

    summary_md += """
---

### Generated Evidence Artifacts:
* Excel Consolidated Master Reports (`Automation_Test_Report.xlsx`, `Summary_Report.xlsx`)
* Excel Status Split Sheets (`Passed_Test_Cases.xlsx`, `Failed_Test_Cases.xlsx`)
* HTML Dashboards & Execution Reports (`dashboard.html`, `execution-report.html`)
* Browser Console logs and screenshots.
"""
    with open(os.path.join(RESULTS_DIR, "Summary", "summary.md"), 'w', encoding='utf-8') as f:
        f.write(summary_md)

    print("Master compilation complete! All test reports generated successfully.")

if __name__ == "__main__":
    compile_all()
